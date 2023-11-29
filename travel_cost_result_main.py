from PyQt5 import QtWidgets, QtCore, QtGui
import travel_cost_result_ui as ui

from datetime import datetime,timedelta
from openpyxl import load_workbook
from openpyxl import Workbook

jp_cityName = {1:"札幌",2:"東京",3:"橫濱",4:"名古屋",5:"京都",
            6:"奈良",7:"大阪",8:"神戶",9:"廣島",10:"沖繩"}
city_search = {1:"札幌",2:"東京都",3:"橫濱市",4:"名古屋",5:"京都府京都市",
            6:"奈良",7:"大阪",8:"神戶",9:"廣島",10:"沖繩島"}
######################################################3
class Main(QtWidgets.QMainWindow , ui.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.lineEdit_PreCost.editingFinished.connect(self.CheckPreCost)
        self.lineEdit_JPplace.editingFinished.connect(self.CheckJPplace)
        self.date_finish.dateChanged.connect(self.CheckDate)
        self.ButtonResult.clicked.connect(self.Check_and_OutPut)
    def CheckJPplace(self):
        catchJPplace = int(self.lineEdit_JPplace.text())
        if catchJPplace<1 and catchJPplace>10:
            self.lineEdit_JPplace.setText("1")
    def CheckPreCost(self):
        catchPreCost = int(self.lineEdit_PreCost.text())
        if catchPreCost<27000 or catchPreCost>32000:
            self.lineEdit_PreCost.setText("27000")
    def CheckDate(self):
        catchSDATE = self.date_start.text()
        catchFDATE = self.date_finish.text()
        Sdate = datetime.strptime(catchSDATE,"%Y%m%d")
        Fdate = datetime.strptime(catchFDATE,"%Y%m%d")
        Dnights = int((Fdate - Sdate).days)
        if Dnights<0 :
            self.label_night_check.setText("日期間距不對")
        if Dnights >= 1:
            self.label_night_check.setText(f"住宿天數：{Dnights}晚")

    def Check_and_OutPut(self):
        #TST = self.lineEdit_PreCost.text()
        CheckSuica = self.checkBox_suica.checkState()
        #print(CheckSuica)##勾選=2，不勾選=0
        ##############[設定值]##########################
        travel_precost = int(self.lineEdit_PreCost.text())
        city_num = int(self.lineEdit_JPplace.text())
        date_start = self.date_start.text()
        date_finish = self.date_finish.text()
        Sdate = datetime.strptime(date_start, "%Y%m%d")
        Fdate = datetime.strptime(date_finish, "%Y%m%d")
        spend_day = int((Fdate - Sdate).days + 1)
        Dnights = int((Fdate - Sdate).days)
        #############[機票]##############
        filenamePlane = date_start + "-" + date_finish + jp_cityName[city_num] + '_航班(來回)列表.xlsx'
        wbPlane = load_workbook(filenamePlane)
        wsPlane = wbPlane[city_search[city_num] + "_航班(來回)"]
        airplane_cost = int(wsPlane.cell(row=2, column=2).value)

        hotel_costs = 0
        hotel_many = 0
        hotel_Dtlist = []
        hotel_Nmlist = []
        hotel_Odlist = []
        hotel_Ntlist = []
        hotel_Rklist = []
        hotel_Cslist = []
        ###########[飯店]#################
        nightLmt = 0
        night_list = [int(self.lineEdit_Hotel1_night.text()),
                      int(self.lineEdit_Hotel2_night.text()),
                      int(self.lineEdit_Hotel3_night.text())]
        print(night_list)

        night_list_et=[]
        for et in range(len(night_list)):
            if night_list[et]!=0:
                night_list_et.append(night_list[et])

        for i in range(len(night_list_et)):
            if nightLmt == 0:
                HTdate = Sdate.strftime("%Y%m%d")
            else:
                HTdate = HTdate.strftime("%Y%m%d")
            print(HTdate)
            hotel_night = night_list[i]
            nights = ''
            if (hotel_night == 1):
                nights = '1晚'
            elif (hotel_night == 2):
                nights = '2晚'
            else:
                break
            print(nights)
            filenameHotel = str(HTdate) + "_" + nights + "_" + jp_cityName[city_num] + '_附近飯店列表(詳細地點).xlsx'
            wbHotel = load_workbook(filenameHotel)
            wsbHotel = wbHotel[jp_cityName[city_num] + "_飯店"]
            hotel_cost = int(wsbHotel.cell(row=2, column=5).value)
            hotel_costs += hotel_cost
            nightLmt += hotel_night
            hotel_many += 1
            #print(nightLmt)
            hotel_Dtlist.append(str(HTdate))
            hotel_Nmlist.append(wsbHotel.cell(row=2, column=1).value)
            hotel_Odlist.append(wsbHotel.cell(row=2, column=6).value)
            hotel_Rklist.append(wsbHotel.cell(row=2, column=3).value)
            hotel_Ntlist.append(str(nights))
            hotel_Cslist.append(str(hotel_cost))
            HTdate = datetime.strptime(HTdate, "%Y%m%d") + timedelta(days=hotel_night)

        # while True:
        #     i = 0  # (night_list位置)
        #     if nightLmt == 0:
        #         HTdate = Sdate.strftime("%Y%m%d")
        #     else:
        #         HTdate = HTdate.strftime("%Y%m%d")
        #     print(HTdate)
        #     hotel_night = night_list[i]
        #     nights = ''
        #     if (hotel_night == 1):
        #         nights = '1晚'
        #     elif (hotel_night == 2):
        #         nights = '2晚'
        #     else:
        #         break
        #     print(nights)
        #     filenameHotel = str(HTdate) + "_" + nights + "_" + jp_cityName[city_num] + '_附近飯店列表(詳細地點).xlsx'
        #     wbHotel = load_workbook(filenameHotel)
        #     wsbHotel = wbHotel[jp_cityName[city_num] + "_飯店"]
        #     hotel_cost = int(wsbHotel.cell(row=2, column=5).value)
        #     hotel_costs += hotel_cost
        #     nightLmt += hotel_night
        #     hotel_many += 1
        #     print(nightLmt)
        #     hotel_Dtlist.append(str(HTdate))
        #     hotel_Nmlist.append(wsbHotel.cell(row=2, column=1).value)
        #     hotel_Odlist.append(wsbHotel.cell(row=2, column=6).value)
        #     hotel_Rklist.append(wsbHotel.cell(row=2, column=3).value)
        #     hotel_Ntlist.append(str(nights))
        #     hotel_Cslist.append(str(hotel_cost))
        #     HTdate = datetime.strptime(HTdate, "%Y%m%d") + timedelta(days=hotel_night)
        #     i += 1
        ###########[景點]#################
        placeTK_costs = 0
        place_many = 0
        place_name_List = []
        placeCs_list = []
        place_OPtime = []

        filenamePlace = jp_cityName[city_num] + '_日本景點列表(詳細地點).xlsx'
        wbPlace = load_workbook(filenamePlace)
        wsPlace = wbPlace[jp_cityName[city_num]]
        take_places = Dnights * 4
        for i in range(2, take_places + 2):
            # print("輸入",i,",",wsPlace.cell(row=i, column=1).value,":門票為",wsPlace.cell(row=i, column=5).value,"元")
            placeTK_costs += int(wsPlace.cell(row=i, column=5).value)
            place_name_List.append(wsPlace.cell(row=i, column=1).value)
            placeCs_list.append(wsPlace.cell(row=i, column=5).value)
            place_OPtime.append(wsPlace.cell(row=i, column=7).value)
            place_many += 1
        ############[其他費用]################
        sim = 500
        suica = 560
        JRroad = 6600
        subway24 = 176
        subway48 = 264
        subway72 = 330

        buy_suica = self.checkBox_suica.checkState()
        if buy_suica == 0:
            suica = 0
        buy_JRroad = self.checkBox_JRroad.checkState()
        if buy_JRroad == 0:
            JRroad = 0
        buy_sub24 = self.checkBox_subway24.checkState()
        if buy_sub24 == 0:
            subway24 = 0
        buy_sub48 = self.checkBox_subway48.checkState()
        if buy_sub48 == 0:
            subway48 = 0
        buy_sub72 = self.checkBox_subway72.checkState()
        if buy_sub72 == 0:
            subway72 = 0
        ################[每天費用]#####################
        meals = int(self.lineEdit_meal_cost.text())
        traffic = int(self.lineEdit_traffic_cost.text())
        joys = int(self.lineEdit_joy_cost.text())
        others = int(self.lineEdit_other_cost.text())

        meals_total = meals * spend_day
        traffic_total = traffic * spend_day
        joys_total = joys * spend_day
        others_total = others * spend_day
        ###############[加總與評估]#######################
        first_cost = airplane_cost + hotel_costs + placeTK_costs
        second_cost = sim + suica + JRroad + subway24 + subway48 + subway72
        other_cost = meals_total + traffic_total + joys_total + others_total

        final_check_cost = first_cost + second_cost + other_cost
        if travel_precost > final_check_cost:
            remain = travel_precost - final_check_cost
            result = f"還有所剩預算{remain}元，可進一步規劃其他有興趣景點，與安排時程"
            self.label_Result_show.setText("結果顯示:"+f"還有所剩預算{remain}元")
        if travel_precost < final_check_cost:
            result = "預算超支，可能要調整預算、擇期選擇機票及(或)飯店(因時間不同，價格會有波動)"
            self.label_Result_show.setText("結果顯示:"+f"還有所剩預算{remain}元"+"，已超支!!")

        wb_travel = Workbook()
        wb_travel.create_sheet("初步規劃")
        ws_travel = wb_travel["初步規劃"]
        your_date=[Sdate,"至",Fdate]
        ws_travel.append(your_date)
        going_to = ["前往日本:",jp_cityName[city_num]]
        ws_travel.append(going_to)
        ws_travel.append([str(spend_day)+"天"+str(Dnights)+"夜"])
        set_total = ["您的預算為:",str(travel_precost),"元(新台幣)"]
        ws_travel.append(set_total)
        ws_travel.append([""])
        ws_travel.append(["機票"])

        plane_top = ["訂票網站", "訂票金額"]
        plane_top1 = ["去程_航空公司",
                     "去程_出發時間", "去程_出發地點", "去程_路徑時間", "去程_抵達時間", "去程_抵達地點", "去程_班次"
                     ]
        plane_top2 = ["回程_航空公司",
                      "回程_出發時間", "回程_出發地點", "回程_路徑時間", "回程_抵達時間", "回程_抵達地點", "回程_班次"
                    ]
        plane0 = []
        plane1 = []
        plane2 = []
        for pln in range(1,wsPlane.max_column+1) :
            if pln==1:
                from_web_txt = str(wsPlane.cell(row=2, column=pln).value).replace("透過","透過爬蟲取得").replace("預訂","、trivago等比價網站")
                plane0.append(from_web_txt)
            if pln==2:
                plane0.append(wsPlane.cell(row=2, column=pln).value)
            if pln>=3 and pln<=9:
                plane1.append(wsPlane.cell(row=2, column=pln).value)
            if pln>9:
                plane2.append(wsPlane.cell(row=2, column=pln).value)
        ws_travel.append(plane_top)
        ws_travel.append(plane0)
        ws_travel.append(plane_top1)
        ws_travel.append(plane1)
        ws_travel.append(plane_top2)
        ws_travel.append(plane2)

        ws_travel.append([""])
        ws_travel.append(["飯店"])
        hotel_top = ["飯店名稱","住宿金額","訂房網","入住日期","住幾晚","評分"]
        ws_travel.append(hotel_top)
        for htl in range(len(hotel_Dtlist)):
            hotel = [hotel_Nmlist[htl],hotel_Cslist[htl],hotel_Odlist[htl],
                     hotel_Dtlist[htl],hotel_Ntlist[htl],hotel_Rklist[htl]]
            ws_travel.append(hotel)
        ws_travel.append(["",hotel_costs])

        ws_travel.append([""])
        ws_travel.append(["景點"])
        place_top = ["景點名稱","門票金額"]
        ws_travel.append(place_top)
        for pce in range(len(place_name_List)):
            place = [place_name_List[pce],placeCs_list[pce]]
            ws_travel.append(place)
        ws_travel.append(["",placeTK_costs])

        ws_travel.append([""])
        ws_travel.append(["網卡、交通卡"])
        ws_travel.append(["網卡",sim])
        ws_travel.append(["suica卡(紅色外國版)", suica])
        ws_travel.append(["JR全日本周遊券", JRroad])
        ws_travel.append(["東京地鐵通票(24小時)", subway24])
        ws_travel.append(["東京地鐵通票(48小時)", subway48])
        ws_travel.append(["東京地鐵通票(72小時)", subway72])
        ws_travel.append([""])

        ws_travel.append(["每天消費現金"])
        ws_travel.append(["","1天的金額",str(spend_day)+"天消費總計"])
        ws_travel.append(["伙食(用餐)費",str(meals), meals_total])
        ws_travel.append(["交通費",str(traffic), traffic_total])
        ws_travel.append(["娛樂費",str(joys), joys_total])
        ws_travel.append(["其他費用",str(others), others_total])
        ws_travel.append(["","",other_cost])

        set_total1 = ["原先預定預算為:", str(travel_precost), "元(新台幣)"]
        ws_travel.append(set_total1)
        ws_travel.append(["費用總計:", final_check_cost, "元(新台幣)"])
        ws_travel.append(["","預定預算，扣除費用總計後"])
        ws_travel.append(["所剩預算:",remain,"元(新台幣)"])

        ws_travel.append([result])
        ws_travel.append(["建議抽空一天為緩衝搭機時間"])
        ws_travel.append(["其他幾天，規劃時程期間建議以早、中、晚安排每天3個景點"])
        savename = jp_cityName[city_num]+"_初步旅遊規劃.xlsx"
        wb_travel.remove(wb_travel['Sheet'])
        wb_travel.save(savename)

if __name__ == '__main__':
    import sys
    app = QtWidgets.QApplication(sys.argv)
    window = Main()
    window.show()
    sys.exit(app.exec_())