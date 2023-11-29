import pandas as pd
from openpyxl import load_workbook
jp_cityName = {1:"札幌",2:"東京",3:"橫濱",4:"名古屋",5:"京都",
            6:"奈良",7:"大阪",8:"神戶",9:"廣島",10:"沖繩"}

from openpyxl import load_workbook
from openpyxl import Workbook

def choose_city():
    city_num = int(input(f"目前有：{jp_cityName}\n請問您要選哪一個都市?(輸入數字)"))
    print("========================================\n")
    return city_num
city = choose_city()
if city>=1 and city<=10:
    cityN = city
filename = jp_cityName[cityN] +'_日本景點列表(詳細地點).xlsx'
wb = load_workbook(filename)
ws=wb[jp_cityName[cityN]]

for cg in range(2,ws.max_row+1):
    if ws.cell(row=cg,column=5).value=="免費" or ws.cell(row=cg,column=5).value=="無顯示門票":
        ws.cell(row=cg, column=5).value = "0"
wb.save(filename)

choose_your_like = int(input("您希望景點能挑有知名度、評價好，還是門票便宜的?\n(輸入數字，1:討論度排序,2:評分排序,3:門票排序)"))
if choose_your_like==1:
    df = pd.read_excel(filename, sheet_name=jp_cityName[cityN])
    wb_value = df.sort_values(by=[ "討論人數", "評分","門票"], ascending=[ False, False,True])
if choose_your_like==2:
    df = pd.read_excel(filename, sheet_name=jp_cityName[cityN])
    wb_value = df.sort_values(by=[ "評分","討論人數", "門票"], ascending=[ False, False,True])
if choose_your_like==3:
    df = pd.read_excel(filename, sheet_name=jp_cityName[cityN])
    wb_value = df.sort_values(by=[ "門票", "評分","討論人數"], ascending=[ True, False,False])

writer = pd.ExcelWriter(filename)
wb_value.to_excel(writer, sheet_name=jp_cityName[cityN], index=False)
writer.save()